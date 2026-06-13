"""
File I/O — streaming export, threading-friendly imports.

* Imports accept an optional ``progress_callback(done, total)``.
* Export streams rows from the DB (no full list in memory).
"""

import csv
import os
import re
from datetime import datetime
from typing import Callable, Optional

from db import Database


# ================================================================ CSV
def import_csv(filepath, db, progress_callback=None):
    warnings: list[str] = []
    with open(filepath, newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        return 0, ["File is empty."]
    parsed = _parse_sheet_rows(rows, warnings)
    return db.bulk_insert(parsed, progress_callback=progress_callback), warnings


# ================================================================ XLSX
def import_xlsx(filepath, db, progress_callback=None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 0, ["openpyxl is required.  Install: pip install openpyxl"]
    warnings: list[str] = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = [
        [("" if c is None else str(c)) for c in row]
        for row in ws.iter_rows(values_only=True)
    ]
    wb.close()
    if not rows:
        return 0, ["Workbook is empty."]
    parsed = _parse_sheet_rows(rows, warnings)
    return db.bulk_insert(parsed, progress_callback=progress_callback), warnings


# ================================================================ ODS
def import_ods(filepath, db, progress_callback=None):
    try:
        from odf.opendocument import load as load_ods
        from odf import table as odf_table
        from odf.teletype import extractText
    except ImportError:
        return 0, ["odfpy is required.  Install: pip install odfpy"]
    warnings: list[str] = []
    doc = load_ods(filepath)
    sheets = doc.spreadsheet.getElementsByType(odf_table.Table)
    if not sheets:
        return 0, ["No sheets found in ODS file."]
    rows: list[list[str]] = []
    for raw_row in sheets[0].getElementsByType(odf_table.TableRow):
        cells = raw_row.getElementsByType(odf_table.TableCell)
        row_vals: list[str] = []
        for cell in cells:
            try:
                repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
            except (TypeError, ValueError):
                repeat = 1
            if repeat > 50:
                repeat = 1
            content = extractText(cell).strip()
            row_vals.extend([content] * repeat)
        rows.append(row_vals)
    if not rows:
        return 0, ["Sheet is empty."]
    parsed = _parse_sheet_rows(rows, warnings)
    return db.bulk_insert(parsed, progress_callback=progress_callback), warnings


# ================================================================ dispatcher
def import_file(filepath, db, progress_callback=None):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        return import_csv(filepath, db, progress_callback)
    if ext in (".xlsx", ".xls"):
        return import_xlsx(filepath, db, progress_callback)
    if ext == ".ods":
        return import_ods(filepath, db, progress_callback)
    return 0, [f"Unsupported file extension: {ext}"]


# ================================================================ export
def export_csv(
    filepath: str, db: Database,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> int:
    """Stream the DB straight into a CSV without materialising a list."""
    total = db.count_entries()
    fmt_hm = db.format_hours_minutes
    written = 0
    with open(filepath, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow([
            "Task Name", "Date", "Start Time", "End Time",
            "Total Minutes", "Rate", "Total",
            "Total Time (H+M)", "Mode", "Notes",
        ])
        for e in db.iter_all_entries():
            writer.writerow([
                e["task_name"], e["date"],
                e["start_time"], e["end_time"],
                e["total_minutes"], e["rate"], e["total"],
                fmt_hm(e["total_minutes"]) if e["mode"] == "hourly" else "",
                e["mode"], e.get("notes", ""),
            ])
            written += 1
            if progress_callback and (written % 250 == 0):
                try:
                    progress_callback(written, total)
                except Exception:
                    pass
    if progress_callback:
        try:
            progress_callback(written, total)
        except Exception:
            pass
    return written


# ================================================================ internals
_SUM_PAT = re.compile(r"^(sum|total|subtotal|grand total|=sum)", re.IGNORECASE)


def _looks_like_header(row):
    return row[0].strip().lower() in (
        "task name", "task", "name", "description", "project", "a", "col a",
    )


def _is_aggregate_row(task_name, row):
    if _SUM_PAT.match(task_name):
        return True
    for idx in (8, 9):
        if idx < len(row) and _SUM_PAT.match(row[idx].strip()):
            return True
    if (not row[2].strip() and not row[3].strip() and not row[5].strip()
            and row[4].strip() and not task_name):
        return True
    return False


def _normalise_date(raw):
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(raw.strip(), fmt).strftime("%m/%d/%y")
        except ValueError:
            continue
    return None


def _parse_number(raw):
    raw = raw.replace("$", "").replace(",", "").strip()
    try:
        return float(raw)
    except (ValueError, TypeError):
        return 0.0


def _parse_sheet_rows(rows, warnings):
    start_idx = 1 if rows and _looks_like_header(rows[0]) else 0
    parsed: list[dict] = []
    for row_num, row in enumerate(rows[start_idx:], start=start_idx + 1):
        while len(row) < 10:
            row.append("")
        task_name = row[0].strip()
        date_str = row[1].strip()
        if not task_name or not date_str:
            continue
        if _is_aggregate_row(task_name, row):
            continue
        date_str = _normalise_date(date_str)
        if date_str is None:
            warnings.append(f"Row {row_num}: unreadable date – skipped.")
            continue
        start_time, end_time = row[2].strip(), row[3].strip()
        has_times = bool(start_time) and bool(end_time)
        mode = "hourly" if has_times else "per_task"
        total_minutes = _parse_number(row[4])
        rate = _parse_number(row[5])
        total = _parse_number(row[6])
        if mode == "hourly" and has_times and total_minutes == 0.0:
            total_minutes = Database.calc_total_minutes(start_time, end_time)
        if total == 0.0 and total_minutes and rate:
            total = Database.calc_total(total_minutes, rate, mode)
        parsed.append({
            "task_name": task_name, "date": date_str,
            "start_time": start_time if has_times else "",
            "end_time": end_time if has_times else "",
            "total_minutes": total_minutes, "rate": rate, "total": total,
            "mode": mode,
            "notes": row[9].strip() if len(row) > 9 else "",
        })
    return parsed